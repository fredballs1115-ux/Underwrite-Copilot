"""Build I-95 Industrial Corridor (Eisenhower Ave Corr + Newington + Springfield + Woodbridge).

Two charts, capped at 2026 (uses '2026 EST' for the 2026 row):
  Sheet 1 - Market vs. Vacancy:      Market Rent $ bars, Vacancy Rate % line
  Sheet 2 - Absorption vs. Deliveries: Net Abs + Deliveries twin bars, Rent Growth line
"""
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

NET_ABS_COLOR   = "963634"
SF_DEL_COLOR    = "C0504D"
RENT_BAR_COLOR  = "963634"
LINE_COLOR      = "1F3864"
GRID_COLOR      = "D9D9D9"

# ---- I-95 Corridor data (CoStar pulls, 2016-2026, '2026 EST' for 2026) ----
DATA = [
    # year, market_rent_psf, vacancy_rate, net_absorption, deliveries (None = no deliveries), rent_growth_yoy
    (2016, 10.27, 0.05386,  436293, None,    0.04101),
    (2017, 10.76, 0.07672, -354654, None,    0.04777),
    (2018, 11.32, 0.06693,  151968, None,    0.05199),
    (2019, 11.99, 0.05288,  398289, 190377,  0.05880),
    (2020, 12.73, 0.04157,  177530, None,    0.06210),
    (2021, 13.99, 0.03434,  113644, None,    0.09832),
    (2022, 15.51, 0.05050, -122441, 138460,  0.10902),
    (2023, 16.83, 0.04980,  225014, 225124,  0.08468),
    (2024, 17.84, 0.06898, -265746,  45600,  0.06011),
    (2025, 18.90, 0.07013,   84595, 110935,  0.05955),
    (2026, 19.59, 0.09715, -306144, 138600,  0.04388),
]

wb = Workbook()

# ============================================================
# SHEET 1: Market Rent vs. Vacancy Rate
# ============================================================
ws = wb.active
ws.title = "Market vs Vacancy"

ws["A1"] = "I-95 Industrial Corridor - Market vs. Vacancy"
ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="000000")
ws.merge_cells("A1:C1")
ws.row_dimensions[1].height = 22

ws["A2"] = "Eisenhower Ave Corr + Newington + Springfield + Woodbridge  |  Annual 2016-2026  |  CoStar"
ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="595959")
ws.merge_cells("A2:C2")

headers1 = ["Year", "Market Rent ($/SF)", "Vacancy Rate"]
for j, h in enumerate(headers1, start=1):
    c = ws.cell(row=4, column=j, value=h)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for i, row in enumerate(DATA, start=5):
    year, rent, vac, _na, _del, _g = row
    cells = [(1, year), (2, rent), (3, vac)]
    for col, val in cells:
        c = ws.cell(row=i, column=col, value=val)
        c.border = border
        c.font = Font(name="Calibri", size=11)
        if col == 1:
            c.alignment = Alignment(horizontal="center")
        elif col == 2:
            c.number_format = '"$"#,##0.00'
            c.alignment = Alignment(horizontal="right")
        elif col == 3:
            c.number_format = "0.00%"
            c.alignment = Alignment(horizontal="right")

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 16

DATA_FIRST = 5
DATA_LAST  = 4 + len(DATA)

bar1 = BarChart()
bar1.type = "col"
bar1.style = 2
bar1.grouping = "clustered"
bar1.gapWidth = 80
bar1.title = "I-95 Industrial Corridor - Market vs. Vacancy"
bar1.y_axis.title = "Market Rent ($/SF)"
bar1.x_axis.title = None

bar1.add_data(Reference(ws, min_col=2, min_row=4, max_row=DATA_LAST), titles_from_data=True)
bar1.set_categories(Reference(ws, min_col=1, min_row=DATA_FIRST, max_row=DATA_LAST))
bar1.series[0].graphicalProperties = GraphicalProperties(solidFill=RENT_BAR_COLOR)
bar1.series[0].graphicalProperties.line = LineProperties(solidFill=RENT_BAR_COLOR)
bar1.y_axis.number_format = '"$"#,##0.00'
bar1.y_axis.majorGridlines.spPr = GraphicalProperties(
    ln=LineProperties(solidFill=GRID_COLOR, w=6350)
)

line1 = LineChart()
line1.add_data(Reference(ws, min_col=3, min_row=4, max_row=DATA_LAST), titles_from_data=True)
line1.y_axis.axId = 200
line1.y_axis.crosses = "max"
line1.y_axis.number_format = "0.0%"
line1.y_axis.title = "Vacancy Rate"
line1.y_axis.majorGridlines = None

s = line1.series[0]
s.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill=LINE_COLOR, w=28000))
s.smooth = False
s.marker = Marker(symbol="circle", size=6)
s.marker.graphicalProperties = GraphicalProperties(solidFill=LINE_COLOR)
s.marker.graphicalProperties.line = LineProperties(solidFill=LINE_COLOR)

bar1 += line1
bar1.height = 11
bar1.width = 22
bar1.legend.position = "b"
ws.add_chart(bar1, "E4")

# ============================================================
# SHEET 2: Absorption vs. Deliveries (twin bars) + Rent Growth line
# ============================================================
ws2 = wb.create_sheet("Absorption vs Deliveries")

ws2["A1"] = "I-95 Industrial Corridor - Absorption vs. Deliveries"
ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="000000")
ws2.merge_cells("A1:D1")
ws2.row_dimensions[1].height = 22

ws2["A2"] = "Eisenhower Ave Corr + Newington + Springfield + Woodbridge  |  Annual 2016-2026  |  CoStar"
ws2["A2"].font = Font(name="Calibri", size=9, italic=True, color="595959")
ws2.merge_cells("A2:D2")

headers2 = ["Year", "Net Absorption", "Deliveries", "Market Rent Growth"]
for j, h in enumerate(headers2, start=1):
    c = ws2.cell(row=4, column=j, value=h)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")

for i, row in enumerate(DATA, start=5):
    year, _rent, _vac, na, deliv, g = row
    cells = [(1, year), (2, na), (3, deliv), (4, g)]
    for col, val in cells:
        c = ws2.cell(row=i, column=col, value=val)
        c.border = border
        c.font = Font(name="Calibri", size=11)
        if col == 1:
            c.alignment = Alignment(horizontal="center")
        elif col in (2, 3):
            c.number_format = '#,##0;[Red](#,##0);"-"'
            c.alignment = Alignment(horizontal="right")
        else:
            c.number_format = "0.00%"
            c.alignment = Alignment(horizontal="right")

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 20

bar2 = BarChart()
bar2.type = "col"
bar2.style = 2
bar2.grouping = "clustered"
bar2.gapWidth = 80
bar2.overlap = -10
bar2.title = "I-95 Industrial Corridor - Absorption vs. Deliveries"
bar2.y_axis.title = "Square Feet"
bar2.x_axis.title = None

bar2.add_data(Reference(ws2, min_col=2, max_col=3, min_row=4, max_row=DATA_LAST), titles_from_data=True)
bar2.set_categories(Reference(ws2, min_col=1, min_row=DATA_FIRST, max_row=DATA_LAST))
bar2.series[0].graphicalProperties = GraphicalProperties(solidFill=NET_ABS_COLOR)
bar2.series[0].graphicalProperties.line = LineProperties(solidFill=NET_ABS_COLOR)
bar2.series[1].graphicalProperties = GraphicalProperties(solidFill=SF_DEL_COLOR)
bar2.series[1].graphicalProperties.line = LineProperties(solidFill=SF_DEL_COLOR)
bar2.y_axis.number_format = '#,##0'
bar2.y_axis.majorGridlines.spPr = GraphicalProperties(
    ln=LineProperties(solidFill=GRID_COLOR, w=6350)
)

line2 = LineChart()
line2.add_data(Reference(ws2, min_col=4, min_row=4, max_row=DATA_LAST), titles_from_data=True)
line2.y_axis.axId = 200
line2.y_axis.crosses = "max"
line2.y_axis.number_format = "0.0%"
line2.y_axis.title = "Market Rent Growth"
line2.y_axis.majorGridlines = None
s2 = line2.series[0]
s2.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill=LINE_COLOR, w=28000))
s2.smooth = False
s2.marker = Marker(symbol="circle", size=6)
s2.marker.graphicalProperties = GraphicalProperties(solidFill=LINE_COLOR)
s2.marker.graphicalProperties.line = LineProperties(solidFill=LINE_COLOR)

bar2 += line2
bar2.height = 11
bar2.width = 22
bar2.legend.position = "b"
ws2.add_chart(bar2, "F4")

out = "/home/user/Underwrite-Copilot/scratchpad/I95_Corridor_Industrial_Charts.xlsx"
wb.save(out)
print(f"Saved: {out}")
