"""Split each two-tab workbook into two single-chart workbooks."""
import shutil
from copy import copy
from openpyxl import load_workbook

PAIRS = [
    ("scratchpad/Woodbridge_Industrial_Chart.xlsx",
     "scratchpad/Woodbridge_1_Market_vs_Vacancy.xlsx",
     "scratchpad/Woodbridge_2_Absorption_vs_Deliveries.xlsx"),
    ("scratchpad/I95_Corridor_Industrial_Charts.xlsx",
     "scratchpad/I95_Corridor_1_Market_vs_Vacancy.xlsx",
     "scratchpad/I95_Corridor_2_Absorption_vs_Deliveries.xlsx"),
]

for src, out1, out2 in PAIRS:
    # Sheet 1 only
    shutil.copy(src, out1)
    wb = load_workbook(out1)
    del wb["Absorption vs Deliveries"]
    wb.save(out1)

    # Sheet 2 only
    shutil.copy(src, out2)
    wb = load_workbook(out2)
    del wb["Market vs Vacancy"]
    wb.save(out2)

    print(f"Split {src}\n  -> {out1}\n  -> {out2}")
