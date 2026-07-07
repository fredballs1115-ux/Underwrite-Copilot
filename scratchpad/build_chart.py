"""Build the Woodbridge industrial submarket charts.

Two charts (2016-2026 only):
  Sheet 1 - Market vs. Vacancy:        Market Rent $ bars + Vacancy Rate line
  Sheet 2 - Absorption vs. Deliveries: Net Abs + SF Delivered twin bars
                                       + Market Rent Growth line
"""
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

NET_ABS_COLOR   = "963634"
SF_DEL_COLOR    = "C0504D"
RENT_BAR_COLOR  = "963634"
LINE_COLOR      = "1F3864"
GRID_COLOR      = "D9D9D9"

# ---- Woodbridge submarket data (CoStar, 2016-2026) ----
DATA = [
    # year, market_rent_psf, vacancy_rate, net_absorption, sf_delivered (None = none), rent_growth_yoy
    (2016,  9.67, 0.02204,  100437, None,    0.0564),
    (2017, 10.15, 0.04520,  -56411, None,    0.0498),
    (2018, 10.74, 0.05503,  -23948, None,    0.0576),
    (2019, 11.37, 0.04733,   18759, None,    0.0585),
    (2020, 12.11, 0.03426,   31836, None,    0.0650),
    (2021, 13.40, 0.02003,   34656, None,    0.1071),
    (2022, 14.93, 0.02389,   -9390, None,    0.1140),
    (2023, 16.19, 0.01949,  231444, 225124,  0.0842),
    (2024, 17.16, 0.03338,  -36955, None,    0.0603),
    (2025, 18.17, 0.03094,  114004, 110935,  0.0591),
    (2026, 18.65, 0.04158,  -29494, None,    0.0591),
]

wb = Workbook()

# ============================================================
# SHEET 1: Market Rent vs. Vacancy Rate
# ============================================================
ws = wb.active
ws.title = "Market vs Vacancy"

ws["A1"] = "Woodbridge Industrial - Market vs. Vacancy"
ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="000000")
ws.merge_cells("A1:C1")
ws.row_dimensions[1].height = 22

ws["A2"] = "Annual 2016-2026  |  CoStar (All Properties -> Industrial, >=20,000 SF)"
ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="595959")
ws.merge_cells("A2:C2")

headers1 = ["Year", "Market Rent ($/SF)", "Vacancy Rate"]
for j, h in enumerate(headers1, start=1):
    c = ws.cell(row=4, column=j, value=h)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for i, row in enumerate(DATA, start=5):
    year, rent, vac, _na, _del, _g = row
    for col, val in [(1, year), (2, rent), (3, vac)]:
        c = ws.cell(row=i, column=col, value=val)
        c.border = border
        c.font = Font(name="Calibri", size=11)
        if col == 1:
            c.alignment = Alignment(horizontal="center")
        elif col == 2:
            c.number_format = '"$"#,##0.00'
            c.alignment = Alignment(horizontal="right")
        elif col == 3:
            c.number_format = "0.00%"
            c.alignment = Alignment(horizontal="right")

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 16

DATA_FIRST = 5
DATA_LAST  = 4 + len(DATA)

bar1 = BarChart()
bar1.type = "col"
bar1.style = 2
bar1.grouping = "clustered"
bar1.gapWidth = 80
bar1.title = "Woodbridge Industrial - Market vs. Vacancy"
bar1.y_axis.title = "Market Rent ($/SF)"
bar1.x_axis.title = None
bar1.add_data(Reference(ws, min_col=2, min_row=4, max_row=DATA_LAST), titles_from_data=True)
bar1.set_categories(Reference(ws, min_col=1, min_row=DATA_FIRST, max_row=DATA_LAST))
bar1.series[0].graphicalProperties = GraphicalProperties(solidFill=RENT_BAR_COLOR)
bar1.series[0].graphicalProperties.line = LineProperties(solidFill=RENT_BAR_COLOR)
bar1.y_axis.number_format = '"$"#,##0.00'
bar1.y_axis.majorGridlines.spPr = GraphicalProperties(
    ln=LineProperties(solidFill=GRID_COLOR, w=6350)
)

line1 = LineChart()
line1.add_data(Reference(ws, min_col=3, min_row=4, max_row=DATA_LAST), titles_from_data=True)
line1.y_axis.axId = 200
line1.y_axis.crosses = "max"
line1.y_axis.number_format = "0.0%"
line1.y_axis.title = "Vacancy Rate"
line1.y_axis.majorGridlines = None
s = line1.series[0]
s.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill=LINE_COLOR, w=28000))
s.smooth = False
s.marker = Marker(symbol="circle", size=6)
s.marker.graphicalProperties = GraphicalProperties(solidFill=LINE_COLOR)
s.marker.graphicalProperties.line = LineProperties(solidFill=LINE_COLOR)

bar1 += line1
bar1.height = 11
bar1.width = 22
bar1.legend.position = "b"
ws.add_chart(bar1, "E4")

# ============================================================
# SHEET 2: Absorption vs. Deliveries + Rent Growth
# ============================================================
ws2 = wb.create_sheet("Absorption vs Deliveries")

ws2["A1"] = "Woodbridge Industrial - Absorption vs. Deliveries"
ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="000000")
ws2.merge_cells("A1:D1")
ws2.row_dimensions[1].height = 22

ws2["A2"] = "Annual 2016-2026  |  CoStar (All Properties -> Industrial, >=20,000 SF)"
ws2["A2"].font = Font(name="Calibri", size=9, italic=True, color="595959")
ws2.merge_cells("A2:D2")

headers2 = ["Year", "Net Absorption", "SF Delivered", "Market Rent Growth"]
for j, h in enumerate(headers2, start=1):
    c = ws2.cell(row=4, column=j, value=h)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")

for i, row in enumerate(DATA, start=5):
    year, _rent, _vac, na, deliv, g = row
    for col, val in [(1, year), (2, na), (3, deliv), (4, g)]:
        c = ws2.cell(row=i, column=col, value=val)
        c.border = border
        c.font = Font(name="Calibri", size=11)
        if col == 1:
            c.alignment = Alignment(horizontal="center")
        elif col in (2, 3):
            c.number_format = '#,##0;[Red](#,##0);"-"'
            c.alignment = Alignment(horizontal="right")
        else:
            c.number_format = "0.00%"
            c.alignment = Alignment(horizontal="right")

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 20

bar2 = BarChart()
bar2.type = "col"
bar2.style = 2
bar2.grouping = "clustered"
bar2.gapWidth = 80
bar2.overlap = -10
bar2.title = "Woodbridge Industrial - Absorption vs. Deliveries"
bar2.y_axis.title = "Square Feet"
bar2.x_axis.title = None
bar2.add_data(Reference(ws2, min_col=2, max_col=3, min_row=4, max_row=DATA_LAST),
              titles_from_data=True)
bar2.set_categories(Reference(ws2, min_col=1, min_row=DATA_FIRST, max_row=DATA_LAST))
bar2.series[0].graphicalProperties = GraphicalProperties(solidFill=NET_ABS_COLOR)
bar2.series[0].graphicalProperties.line = LineProperties(solidFill=NET_ABS_COLOR)
bar2.series[1].graphicalProperties = GraphicalProperties(solidFill=SF_DEL_COLOR)
bar2.series[1].graphicalProperties.line = LineProperties(solidFill=SF_DEL_COLOR)
bar2.y_axis.number_format = '#,##0'
bar2.y_axis.majorGridlines.spPr = GraphicalProperties(
    ln=LineProperties(solidFill=GRID_COLOR, w=6350)
)

line2 = LineChart()
line2.add_data(Reference(ws2, min_col=4, min_row=4, max_row=DATA_LAST), titles_from_data=True)
line2.y_axis.axId = 200
line2.y_axis.crosses = "max"
line2.y_axis.number_format = "0.0%"
line2.y_axis.title = "Market Rent Growth"
line2.y_axis.majorGridlines = None
s2 = line2.series[0]
s2.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill=LINE_COLOR, w=28000))
s2.smooth = False
s2.marker = Marker(symbol="circle", size=6)
s2.marker.graphicalProperties = GraphicalProperties(solidFill=LINE_COLOR)
s2.marker.graphicalProperties.line = LineProperties(solidFill=LINE_COLOR)

bar2 += line2
bar2.height = 11
bar2.width = 22
bar2.legend.position = "b"
ws2.add_chart(bar2, "F4")

out = "/home/user/Underwrite-Copilot/scratchpad/Woodbridge_Industrial_Chart.xlsx"
wb.save(out)
print(f"Saved: {out}")
