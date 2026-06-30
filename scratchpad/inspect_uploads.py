"""Dump the I-95 Corridor CoStar pulls so I can see structure + values."""
from openpyxl import load_workbook
from pathlib import Path

UP = Path("/root/.claude/uploads/262966f4-a9e4-5739-b6f8-4a328b7b1602")
files = [
    "057a99a9-Market_Asking_Rent_Per_SF_1.xlsx",
    "d5e008af-Market_Asking_Rent_Growth_YOY.xlsx",
    "90323ff7-Vacancy_Rate_2.xlsx",
    "c383dd19-Net_Absorption_1.xlsx",
    "d281982a-Deliveries.xlsx",
]

for f in files:
    p = UP / f
    print("\n" + "=" * 80)
    print(f"FILE: {f}")
    print("=" * 80)
    wb = load_workbook(p, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        print(f"  Sheet: {sh!r}  rows={ws.max_row}  cols={ws.max_column}")
        for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True):
            print("   ", r)
