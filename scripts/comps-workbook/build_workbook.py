#!/usr/bin/env python3
"""Assemble a master comps workbook from per-package extraction JSON files.

Each JSON file describes one sales package (offering memorandum) and the
lease / sale / land-sale comps extracted from it.  This script merges them,
dedupes comps that appear in multiple packages, buckets rows by the comp
property's state, and writes one xlsx with a tab per (state x comp type)
plus a Sources tab logging every package processed.

Usage:
    python3 build_workbook.py --input-dir <dir-of-json> --output <path.xlsx> \
        [--states VA MD PA] [--title "2026 Industrial Comps"]

Expected JSON shape per file (missing keys / nulls are fine):
{
  "package": {"file", "offering_name", "property_address", "city", "state",
               "property_type", "broker", "pages"},
  "lease_comps": [...], "sale_comps": [...], "land_sale_comps": [...],
  "sections_found": [...], "coverage": "...", "warnings": [...]
}
"""

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LEASE_COLUMNS = [
    ("property", "Property"),
    ("address", "Address"),
    ("city", "City"),
    ("state", "State"),
    ("submarket", "Submarket"),
    ("tenant", "Tenant"),
    ("leased_sf", "Leased SF"),
    ("rate", "Rate (as quoted)"),
    ("rate_basis", "Rate Basis"),
    ("rate_psf_yr", "Rate $/SF/Yr"),
    ("lease_structure", "Structure"),
    ("escalations", "Escalations"),
    ("term", "Term"),
    ("lease_date", "Lease Date"),
    ("notes", "Notes"),
    ("broker", "Broker"),
    ("source", "Source Package(s)"),
    ("page_ref", "PDF Page(s)"),
]

SALE_COLUMNS = [
    ("property", "Property"),
    ("address", "Address"),
    ("city", "City"),
    ("state", "State"),
    ("building_sf", "Building SF"),
    ("year_built", "Year Built"),
    ("sale_price", "Sale Price"),
    ("price_psf", "Price $/SF"),
    ("cap_rate", "Cap Rate"),
    ("occupancy", "Occupancy"),
    ("buyer", "Buyer"),
    ("seller", "Seller"),
    ("sale_date", "Sale Date"),
    ("notes", "Notes"),
    ("broker", "Broker"),
    ("source", "Source Package(s)"),
    ("page_ref", "PDF Page(s)"),
]

LAND_COLUMNS = [
    ("property", "Property / Location"),
    ("address", "Address"),
    ("city", "City"),
    ("state", "State"),
    ("acres", "Acres"),
    ("zoning", "Zoning"),
    ("sale_price", "Sale Price"),
    ("price_per_acre", "Price $/Acre"),
    ("intended_use", "Intended Use"),
    ("buyer", "Buyer"),
    ("seller", "Seller"),
    ("sale_date", "Sale Date"),
    ("notes", "Notes"),
    ("broker", "Broker"),
    ("source", "Source Package(s)"),
    ("page_ref", "PDF Page(s)"),
]

COMP_TYPES = [
    ("lease_comps", "Lease Comps", LEASE_COLUMNS),
    ("sale_comps", "Sales Comps", SALE_COLUMNS),
    ("land_sale_comps", "Land Sale Comps", LAND_COLUMNS),
]

LAND_LEASE_COLUMNS = [
    ("property", "Property / Location"),
    ("address", "Address"),
    ("city", "City"),
    ("state", "State"),
    ("tenant", "Tenant"),
    ("landlord", "Landlord / Owner"),
    ("acres", "Acres"),
    ("rate_mo_acre", "Rate $/Mo/Acre"),
    ("term", "Term"),
    ("lease_date", "Lease Date"),
    ("notes", "Notes"),
    ("broker", "Broker"),
    ("source", "Source Package(s)"),
    ("page_ref", "PDF Page(s)"),
]

# Comp categories outside the per-state tab set; each gets one combined tab.
EXTRA_TYPES = [
    ("office_sale_comps", "DMV Office Sale Comps", SALE_COLUMNS),
    ("land_lease_comps", "Land Lease Comps (IOS)", LAND_LEASE_COLUMNS),
]

NUMBER_FORMATS = {
    "leased_sf": "#,##0",
    "building_sf": "#,##0",
    "sale_price": "$#,##0",
    "price_psf": "$#,##0.00",
    "price_per_acre": "$#,##0",
    "rate_psf_yr": "$#,##0.00",
    "acres": "#,##0.00",
}

STATE_NAMES = {
    "virginia": "VA", "maryland": "MD", "pennsylvania": "PA",
    "district of columbia": "DC", "washington dc": "DC", "d.c.": "DC",
    "delaware": "DE", "new jersey": "NJ", "west virginia": "WV",
    "north carolina": "NC", "new york": "NY",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(size=10)

BROKER_NAMES = [
    ("cushman", "Cushman & Wakefield"), ("c&w", "Cushman & Wakefield"),
    ("newmark", "Newmark"), ("nmrk", "Newmark"), ("jll", "JLL"),
    ("jones lang", "JLL"), ("cbre", "CBRE"), ("berkadia", "Berkadia"),
    ("transwestern", "Transwestern"), ("klnb", "KLNB"),
    ("nai michael", "NAI Michael"), ("jm zell", "JM Zell"),
    ("kbc", "KBC Advisors"), ("lincoln", "Lincoln Property Company"),
]


def broker_short(value):
    """Reduce a broker credit line to just the company name."""
    if not value:
        return None
    low = str(value).lower()
    for needle, name in BROKER_NAMES:
        if needle in low:
            return name
    return str(value).split("(")[0].split(",")[0].split("|")[0].strip()


# (pattern in note text, terse label) — order matters, first match groups first.
NOTE_FLAGS = [
    (r"subject propert|subject portfolio", "Subject property"),
    (r"under contract|\bu/c\b", "Under contract"),
    (r"\bpending\b", "Pending"),
    (r"not an executed|not executed|asking rate|availabilit|marketing", "Asking, not executed"),
    (r"pipeline|target rent", "Pipeline target rent"),
    (r"discrepancy", "Source discrepancy"),
    (r"converted from monthly", "Converted from monthly"),
    (r"sale.?leaseback", "Sale-leaseback"),
    (r"\breo\b|deed.in.lieu|note sale", "Distressed/note sale"),
    (r"vacant", "Vacant at sale"),
    (r"owner.user|user buyer|user purchase", "Owner-user sale"),
    (r"portfolio", "Portfolio deal"),
    (r"renewal", "Renewal"),
    (r"new lease", "New lease"),
    (r"expansion", "Expansion"),
    (r"land value|razed|redevelop", "Land-value trade"),
    (r"forward take|takeout|earn.out", "Forward/earn-out"),
    (r"stabilized", "Stabilized cap"),
    (r"\broc\b|return on cost", "RoC, not cap"),
    (r"20(21|22|23)", "Pre-2024 vintage"),
]


def shorten_note(text):
    """Compress a long note into a few flag words for the workbook.
    Full detail stays in the extraction JSONs."""
    if not text:
        return None
    low = str(text).lower()
    flags = []
    for pattern, label in NOTE_FLAGS:
        if re.search(pattern, low) and label not in flags:
            flags.append(label)
        if len(flags) == 3:
            break
    if flags:
        return "; ".join(flags)
    first = re.split(r"[;.]", str(text))[0].strip()
    return (first[:57] + "...") if len(first) > 60 else (first or None)


WARNING_KEEP = [
    (r"truncat|missing from this file|ends at", "File truncated — section(s) missing"),
    (r"qc pass corrected|transcription error", "QC corrected values vs source"),
    (r"2022", "2022 vintage"),
    (r"2025-vintage|months of 2025|2025 comp", "2025 vintage"),
    (r"duplicate", "Duplicate upload"),
    (r"under contract|pending", "Some rows pending/U-C"),
    (r"asking|availabilit|pipeline|not executed", "Some rows asking, not executed"),
    (r"misprint", "Source misprint (value nulled)"),
    (r"lab/life science|residential|office", "Non-industrial rows removed"),
]


def shorten_basis(value):
    """Compress verbose rate-basis strings to a terse label."""
    if not value:
        return None
    low = str(value).lower()
    if "nnn" in low:
        return "$/SF NNN"
    if "/mo" in low or "per month" in low or ("monthly" in low and "annual" not in low):
        return "$/SF/mo"
    if "gross" in low or re.search(r"\bfs\b", low):
        return "Gross/FS"
    if "$/sf" in low or "psf" in low or "per sf" in low:
        return "$/SF"
    if "not stated" in low or "not printed" in low:
        return "Not stated"
    return str(value)[:20]


ROW_LEVEL_WARNINGS = {"Some rows pending/U-C", "Some rows asking, not executed",
                      "QC corrected values vs source", "Non-industrial rows removed"}


def shorten_warnings(warnings, has_comps=True):
    """Keep only genuinely notable warnings, as short phrases.
    Row-level flags are meaningless for packages that contributed no comps."""
    out = []
    low_all = " ".join(str(w).lower() for w in warnings or [])
    for pattern, label in WARNING_KEEP:
        if label in ROW_LEVEL_WARNINGS and not has_comps:
            continue
        if re.search(pattern, low_all) and label not in out:
            out.append(label)
    return "; ".join(out)


def norm_state(value):
    if not value:
        return None
    s = str(value).strip()
    if len(s) == 2:
        return s.upper()
    return STATE_NAMES.get(s.lower(), s.upper())


MONTH_ABBR = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}
MONTH_FULL = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July", "August",
     "September", "October", "November", "December"], 1)}


def norm_month(value):
    """Normalize a printed date to 'YYYY-MM' when parseable, else None."""
    if not value:
        return None
    s = str(value).strip().lower()
    m = re.match(r"([a-z]+)[\s\-/]+(\d{2,4})", s)
    if m and (m.group(1) in MONTH_ABBR or m.group(1) in MONTH_FULL):
        month = MONTH_ABBR.get(m.group(1)) or MONTH_FULL.get(m.group(1))
        yr = int(m.group(2))
        return f"{yr if yr > 99 else 2000 + yr}-{month:02d}"
    m = re.match(r"(\d{4})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.match(r"(\d{1,2})/\d{1,2}/(\d{4})", s)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    return None


def dedupe_key(comp, comp_type):
    """Identity of a comp across packages.

    Address alone is not identity: two tenants can lease the same building and
    a building can trade twice, so leases key on tenant as well and sales on
    the transaction month. A same-key value conflict is split later by the
    guard in collect().
    """
    base = comp.get("address") or comp.get("property") or ""
    city = comp.get("city") or ""
    state = norm_state(comp.get("state")) or ""
    parts = [base, city, state]
    if comp_type in ("lease_comps", "land_lease_comps"):
        parts.append(comp.get("tenant") or "")
    else:
        parts.append(norm_month(comp.get("sale_date")) or "")
    key = re.sub(r"[^a-z0-9]", "", "".join(str(p) for p in parts).lower())
    return key or None


def guard_conflict(existing, incoming, comp_type):
    """True when two same-key comps are actually distinct transactions:
    their headline numbers differ by more than tolerance."""
    if comp_type in ("lease_comps", "land_lease_comps"):
        checks = [("leased_sf", 0.02), ("rate", 0.05), ("rate_mo_acre", 0.05)]
    else:
        checks = [("sale_price", 0.05)]
    for field, tol in checks:
        a, b = existing.get(field), incoming.get(field)
        if isinstance(a, (int, float)) and isinstance(b, (int, float)) and a and b:
            if abs(a - b) / max(abs(a), abs(b)) > tol:
                return True
    return False


def merge_comps(existing, incoming, source):
    """Fold an incoming duplicate into the existing row; flag value conflicts."""
    conflicts = []
    for k, v in incoming.items():
        if k in ("pages", "notes", "source", "page_ref") or v in (None, ""):
            continue
        cur = existing.get(k)
        if cur in (None, ""):
            existing[k] = v
        elif str(cur).strip().lower() != str(v).strip().lower():
            conflicts.append(f"{k}: {source} shows '{v}'")
    if incoming.get("notes") and incoming["notes"] != existing.get("notes"):
        existing["notes"] = "; ".join(filter(None, [existing.get("notes"), incoming["notes"]]))
    if conflicts:
        flag = "DISCREPANCY " + "; ".join(conflicts)
        existing["notes"] = "; ".join(filter(None, [existing.get("notes"), flag]))


def load_packages(input_dir):
    packages = []
    for path in sorted(Path(input_dir).glob("*.json")):
        with open(path) as f:
            data = json.load(f)
        data.setdefault("package", {})
        data["package"].setdefault("file", path.name)
        packages.append(data)
    return packages


def collect(packages):
    """Return {comp_type: {key or unique id: row}} with sources merged."""
    collected = {ct: {} for ct, _, _ in COMP_TYPES + EXTRA_TYPES}
    counter = 0
    for pkg in packages:
        source = pkg["package"].get("offering_name") or pkg["package"]["file"]
        broker = pkg["package"].get("broker")
        for ct, _, _ in COMP_TYPES + EXTRA_TYPES:
            for comp in pkg.get(ct) or []:
                counter += 1
                comp = dict(comp)
                comp["state"] = norm_state(comp.get("state"))
                pages = comp.pop("pages", None) or []
                page_str = ", ".join(f"p.{p}" for p in pages)
                key = dedupe_key(comp, ct) or f"__uniq{counter}"
                bucket = collected[ct]
                while key in bucket and guard_conflict(bucket[key], comp, ct):
                    key += "~"  # same key but conflicting numbers: a distinct transaction
                if key in bucket:
                    row = bucket[key]
                    merge_comps(row, comp, source)
                    row["_sources"].append((source, page_str, broker))
                else:
                    comp["_sources"] = [(source, page_str, broker)]
                    bucket[key] = comp
    return collected


def write_sheet(wb, title, columns, rows):
    ws = wb.create_sheet(title)
    for col_idx, (_, header) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    widths = [len(h) for _, h in columns]
    for row_idx, row in enumerate(rows, start=2):
        sources = row.get("_sources", [])
        row = dict(row)
        row["source"] = "; ".join(dict.fromkeys(s for s, _, _ in sources))
        row["page_ref"] = "; ".join(p for _, p, _ in sources if p)
        row["broker"] = "; ".join(dict.fromkeys(broker_short(b) for _, _, b in sources if b))
        row["notes"] = shorten_note(row.get("notes"))
        row["rate_basis"] = shorten_basis(row.get("rate_basis"))
        if not row.get("property"):
            row["property"] = row.get("address")
        for col_idx, (field, _) in enumerate(columns, start=1):
            value = row.get(field)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            if field in NUMBER_FORMATS and isinstance(value, (int, float)):
                cell.number_format = NUMBER_FORMATS[field]
            if value is not None:
                widths[col_idx - 1] = max(widths[col_idx - 1], min(len(str(value)), 32))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w + 3
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    return ws


def write_sources_sheet(wb, packages, collected):
    columns = ["Offering", "Address", "City", "State", "Type",
               "Broker", "Pages", "Lease Comps", "Sale Comps", "Land Comps",
               "Comps Found", "Warnings"]
    ws = wb.create_sheet("Sources")
    for i, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    packages = sorted(packages, key=lambda pkg: str(pkg["package"].get("city") or "~"))
    for r, pkg in enumerate(packages, start=2):
        p = pkg["package"]
        n_comps = sum(len(pkg.get(ct) or []) for ct, _, _ in COMP_TYPES + EXTRA_TYPES)
        values = [
            p.get("offering_name") or p.get("file"), p.get("property_address"),
            p.get("city"), norm_state(p.get("state")), "Industrial",
            broker_short(p.get("broker")), p.get("pages"),
            len(pkg.get("lease_comps") or []),
            len(pkg.get("sale_comps") or []),
            len(pkg.get("land_sale_comps") or []),
            "Yes" if n_comps else "No",
            shorten_warnings(pkg.get("warnings"), has_comps=n_comps > 0),
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=v).font = BODY_FONT
    ws.freeze_panes = "A2"
    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-dir", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--states", nargs="+", default=["VA", "MD", "PA"])
    ap.add_argument("--title", default="Comps")
    args = ap.parse_args()

    packages = load_packages(args.input_dir)
    if not packages:
        raise SystemExit(f"No extraction JSON files found in {args.input_dir}")
    collected = collect(packages)

    wb = Workbook()
    wb.remove(wb.active)
    tab_counts = {}
    leftovers = {}
    for state in args.states:
        for ct, label, columns in COMP_TYPES:
            rows = [c for c in collected[ct].values() if c.get("state") == state]
            rows.sort(key=lambda c: (str(c.get("city") or "~").lower(), str(c.get("property") or c.get("address") or "").lower()))
            name = f"{state} {label}"
            write_sheet(wb, name, columns, rows)
            tab_counts[name] = len(rows)
    for ct, label, columns in COMP_TYPES:
        rows = [c for c in collected[ct].values() if c.get("state") not in args.states]
        if rows:
            leftovers[label] = rows
    if leftovers:
        for label, rows in leftovers.items():
            columns = next(cols for _, l, cols in COMP_TYPES if l == label)
            rows.sort(key=lambda c: (str(c.get("city") or "~").lower(), str(c.get("property") or c.get("address") or "").lower()))
            name = f"Other States {label}"[:31]
            write_sheet(wb, name, columns, rows)
            tab_counts[name] = len(rows)
    for ct, label, columns in EXTRA_TYPES:
        rows = list(collected[ct].values())
        if rows:
            rows.sort(key=lambda c: (str(c.get("city") or "~").lower(), str(c.get("property") or c.get("address") or "").lower()))
            write_sheet(wb, label[:31], columns, rows)
            tab_counts[label] = len(rows)
    write_sources_sheet(wb, packages, collected)

    wb.save(args.output)
    print(f"Wrote {args.output}")
    for name, n in tab_counts.items():
        print(f"  {name}: {n}")
    print(f"  Sources: {len(packages)} packages")


if __name__ == "__main__":
    main()
